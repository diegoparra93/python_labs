import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Converts CSV to XLSX using openpyxl
    """
    print(f"Starting conversion: {csv_path} → {xlsx_path}")

    # Check if the file exists
    csv_file = Path(csv_path)
    if not csv_file.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_path}")

    # Read CSV file
    with open(csv_path, "r", encoding="utf-8") as cf:
        reader = csv.reader(cf)
        rows = list(reader)

    print(f"Rows read from CSV: {len(rows)}")

    if not rows:
        raise ValueError("CSV file is empty")

    # Show content for debugging
    for i, row in enumerate(rows):
        print(f"  Row {i}: {row}")

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Write data to Excel
    for row in rows:
        ws.append(row)

    print("Data written to XLSX")

    # Adjust auto-width for columns
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = max(max_length + 2, 8)
        ws.column_dimensions[column_letter].width = adjusted_width
        print(f"  Column {column_letter}: width {adjusted_width}")

    # Create directory if it doesn't exist
    Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)

    # Save the file (THIS IS THE IMPORTANT PART!)
    wb.save(xlsx_path)
    print(f"XLSX saved successfully: {xlsx_path}")

    # Verify that the file was created
    if Path(xlsx_path).exists():
        file_size = Path(xlsx_path).stat().st_size
        print(f"File size: {file_size} bytes")
    else:
        print("ERROR: XLSX file was not created")


# Example of how to use the function:
if __name__ == "__main__":
    try:
        csv_to_xlsx("data/samples/people.csv", "data/out/people.xlsx")
        print("Conversion completed successfully!")
    except Exception as e:
        print(f"Error during conversion: {e}")
